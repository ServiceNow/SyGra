"""
Generate Excel with metrics for all judge model runs across both tasks.
Uses each task's own metrics.py for computation.
"""

import importlib.util
import os
import sys
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Load task-specific metrics modules ─────────────────────────────────────────

def load_metrics_module(task_dir):
    path = os.path.join(task_dir, "metrics.py")
    spec = importlib.util.spec_from_file_location("metrics", path)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


SYNTHESIS_DIR = "tasks/examples/search_qna_synthesis"
ROUTER_DIR    = "tasks/examples/search_qna_router_query_rewrite"

synthesis_metrics = load_metrics_module(SYNTHESIS_DIR)
router_metrics    = load_metrics_module(ROUTER_DIR)

# ── Task configs ───────────────────────────────────────────────────────────────

SYNTHESIS_CLASSES = ["Direct Answer", "Clarification Question", "No Answer Found", "Either Direct or Clarification"]
ROUTER_CLASSES    = ["va", "question_answering", "small_talk_generic", "small_talk_closure", "small_talk_displeasure", "agent"]

SYNTHESIS_FILES = {
    "gpt-5":           f"{SYNTHESIS_DIR}/GPT5.json",
    "smollm3":         f"{SYNTHESIS_DIR}/smollm3_fixed_output_2026-03-25_17-10-40.json",
    "ministral":       f"{SYNTHESIS_DIR}/ministral_fixed_output_2026-03-25_17-10-40.json",
    "granite":         f"{SYNTHESIS_DIR}/granite_fixed_output_2026-03-25_17-10-40.json",
    "olmo":            f"{SYNTHESIS_DIR}/olmo_fixed_output_2026-03-25_17-10-40.json",
    "jan_release_rc8": f"{SYNTHESIS_DIR}/jan_release_rc8_fixed_output_2026-03-25_17-47-45.json",
    "gemma":           f"{SYNTHESIS_DIR}/gemma_fixed_full_output_2026-03-25_17-51-58.json",
}

ROUTER_FILES = {
    "gpt-5":           f"{ROUTER_DIR}/gpt-5_output_2026-03-24_21-59-02.json",
    "smollm3":         f"{ROUTER_DIR}/smollm3_router_fixed_output_2026-03-25_17-26-36.json",
    "ministral":       f"{ROUTER_DIR}/ministral_router_fixed_output_2026-03-25_17-26-50.json",
    "granite":         f"{ROUTER_DIR}/granite_router_fixed_output_2026-03-25_17-26-37.json",
    "olmo":            f"{ROUTER_DIR}/olmo_router_fixed_output_2026-03-25_17-26-35.json",
    "jan_release_rc8": f"{ROUTER_DIR}/jan_release_rc8_router_fixed_output_2026-03-25_17-47-58.json",
    "gemma":           f"{ROUTER_DIR}/gemma_router_fixed_output_2026-03-25_17-26-34.json",
}

# ── Excel helpers ──────────────────────────────────────────────────────────────

HEADER_FILL    = PatternFill("solid", fgColor="1F4E79")
SUBHEADER_FILL = PatternFill("solid", fgColor="2E75B6")
MACRO_FILL     = PatternFill("solid", fgColor="D6E4F0")
ALT_FILL       = PatternFill("solid", fgColor="EBF3FB")
WHITE_FILL     = PatternFill("solid", fgColor="FFFFFF")
PARTIAL_FILL   = PatternFill("solid", fgColor="FFF2CC")

HEADER_FONT    = Font(bold=True, color="FFFFFF", size=11)
SUBHEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
MACRO_FONT     = Font(bold=True, size=10)
CENTER         = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT           = Alignment(horizontal="left",   vertical="center")

thin   = Side(style="thin", color="AAAAAA")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

METRICS = ["precision", "recall", "f1", "accuracy", "support", "TP", "FP", "TN", "FN"]
METRIC_HEADERS = ["Precision", "Recall", "F1", "Accuracy", "Support", "TP", "FP", "TN", "FN"]


def style_cell(cell, fill=None, font=None, alignment=None):
    if fill:      cell.fill = fill
    if font:      cell.font = font
    if alignment: cell.alignment = alignment
    cell.border = BORDER


def write_task_sheet(wb, sheet_name, files, classes, metrics_mod, full_count):
    ws = wb.create_sheet(sheet_name)

    # Collect all model metrics using the task's own metrics.py
    all_model_metrics = {}
    for model, fpath in files.items():
        if not os.path.exists(fpath):
            print(f"  [SKIP] {model}: file not found")
            continue
        data = metrics_mod.load(fpath)
        n = len(data)
        cm = metrics_mod.build_confusion(data)
        m  = metrics_mod.compute_metrics(cm)
        all_model_metrics[model] = (m, n)
        print(f"  {model}: {n} records")

    models     = list(all_model_metrics.keys())
    all_classes = classes + ["macro_avg"]

    # ── Header row 1: sheet title ─────────────────────────────────────────────
    row = 1
    ws.cell(row, 1, sheet_name).font = Font(bold=True, size=13)
    ws.cell(row, 1).alignment = CENTER
    row += 1

    # ── Header row 2: model names spanning metric columns ────────────────────
    ws.cell(row, 1, "Class").fill = HEADER_FILL
    ws.cell(row, 1).font = HEADER_FONT
    ws.cell(row, 1).alignment = CENTER
    col = 2
    for model in models:
        _, n = all_model_metrics[model]
        label = f"{model}\n({n} records)"
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + len(METRICS) - 1)
        c = ws.cell(row, col, label)
        is_partial = n < full_count
        fill = PARTIAL_FILL if is_partial else SUBHEADER_FILL
        font = Font(bold=True, size=10, color="7F4700") if is_partial else SUBHEADER_FONT
        style_cell(c, fill=fill, font=font, alignment=CENTER)
        col += len(METRICS)
    row += 1

    # ── Header row 3: metric names ────────────────────────────────────────────
    ws.cell(row, 1, "Class").fill = HEADER_FILL
    ws.cell(row, 1).font = HEADER_FONT
    ws.cell(row, 1).alignment = CENTER
    col = 2
    for model in models:
        for header in METRIC_HEADERS:
            c = ws.cell(row, col, header)
            style_cell(c, fill=SUBHEADER_FILL, font=SUBHEADER_FONT, alignment=CENTER)
            col += 1
    row += 1

    # ── Data rows ─────────────────────────────────────────────────────────────
    for i, cls in enumerate(all_classes):
        is_macro  = cls == "macro_avg"
        row_fill  = MACRO_FILL if is_macro else (ALT_FILL if i % 2 == 0 else WHITE_FILL)
        row_font  = MACRO_FONT if is_macro else Font(size=10)

        label = "MACRO AVG" if is_macro else cls
        c = ws.cell(row, 1, label)
        style_cell(c, fill=row_fill, font=Font(bold=is_macro, size=10), alignment=LEFT)

        col = 2
        for model in models:
            m, _ = all_model_metrics[model]
            cls_data = m.get(cls, {})
            for key in METRICS:
                val = cls_data.get(key, "N/A")
                if isinstance(val, float):
                    val = round(val, 4)
                c = ws.cell(row, col, val)
                style_cell(c, fill=row_fill, font=row_font, alignment=CENTER)
                col += 1
        row += 1

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 30
    col = 2
    for model in models:
        for _ in METRICS:
            ws.column_dimensions[get_column_letter(col)].width = 11
            col += 1

    ws.row_dimensions[2].height = 35
    ws.row_dimensions[3].height = 20
    ws.freeze_panes = "B4"


# ── Main ───────────────────────────────────────────────────────────────────────

wb = openpyxl.Workbook()
wb.remove(wb.active)

print("=== Synthesis ===")
write_task_sheet(
    wb, "Synthesis (QnA)",
    SYNTHESIS_FILES,
    SYNTHESIS_CLASSES,
    metrics_mod=synthesis_metrics,
    full_count=400,
)

print("=== Router ===")
write_task_sheet(
    wb, "Router (Query Rewrite)",
    ROUTER_FILES,
    ROUTER_CLASSES,
    metrics_mod=router_metrics,
    full_count=2700,
)

out = "judge_metrics_all_models.xlsx"
wb.save(out)
print(f"\nSaved: {out}")
